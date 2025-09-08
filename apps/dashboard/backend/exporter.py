# apps/dashboard/backend/exporter.py
from __future__ import annotations
from io import BytesIO
from typing import Dict, Any, List
from datetime import datetime
from pathlib import Path
import unicodedata
import os

def _nowstamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

def _maybe_save_copy(data: bytes, run_id: str, ext: str) -> None:
    out_root = os.getenv("SECAPP_REPORT_DIR")
    if not out_root:
        return
    d = os.path.join(out_root, run_id)
    os.makedirs(d, exist_ok=True)
    name = f"security-app_{run_id}_{_nowstamp()}.{ext}"
    with open(os.path.join(d, name), "wb") as f:
        f.write(data)

# ---------- NEW: Unicode font helpers ----------
_UNI_REG_CANDS = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
    os.path.join(os.path.dirname(__file__), "fonts", "DejaVuSans.ttf"),
]
_UNI_BOLD_CANDS = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf",
    os.path.join(os.path.dirname(__file__), "fonts", "DejaVuSans-Bold.ttf"),
]

def _pick_font(cands: list[str]) -> str | None:
    for p in cands:
        if Path(p).exists():
            return p
    return None

def _ascii_only(s: Any) -> str:
    # thay em dash bằng dấu gạch thường, loại bỏ ký tự ngoài Latin-1
    t = str(s).replace("—", "-")
    return unicodedata.normalize("NFKD", t).encode("latin-1", "ignore").decode("latin-1")

def _setup_unicode(pdf) -> bool:
    reg = _pick_font(_UNI_REG_CANDS)
    if not reg:
        return False
    pdf.add_font("U", "", reg, uni=True)
    bold = _pick_font(_UNI_BOLD_CANDS)
    if bold:
        pdf.add_font("U", "B", bold, uni=True)
    pdf.set_font("U", size=12)
    return True
# ---------- /NEW ----------

def build_excel(summary: Dict[str, Any], rules: List[Dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active; ws.title = "Summary"
    for k in ("total_rules","all_ok","with_failures","pass_rate",
              "total_commands","commands_ok","commands_failed"):
        ws.append([k, summary.get(k)])

    ws2 = wb.create_sheet("By severity")
    ws2.append(["severity","rules","rules_ok","cmd_ok","cmd_fail"])
    for sev, d in (summary.get("by_severity") or {}).items():
        ws2.append([sev, d.get("rules",0), d.get("rules_ok",0), d.get("cmd_ok",0), d.get("cmd_fail",0)])

    ws3 = wb.create_sheet("Top failing")
    ws3.append(["#","id","severity","title","cmd_ok","cmd_fail","status"])
    for i, t in enumerate(summary.get("top_failing_rules") or [], start=1):
        ws3.append([i, t.get("id"), t.get("severity"), t.get("title"), t.get("cmd_ok"), t.get("cmd_fail"), t.get("status")])

    ws4 = wb.create_sheet("Denied")
    ws4.append(["id","severity","title","#denied","examples"])
    for r in summary.get("denied_rules") or []:
        ws4.append([r.get("id"), r.get("severity"), r.get("title"), r.get("denied"), ", ".join(r.get("examples", []))])

    ws5 = wb.create_sheet("Rules")
    ws5.append(["id","severity","title","cmd_ok","cmd_fail","status"])
    for r in rules or []:
        ws5.append([r.get("id"), r.get("severity"), r.get("title"), r.get("cmd_ok"), r.get("cmd_fail"), r.get("status")])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def build_pdf(run_id: str, summary: Dict[str, Any]) -> bytes:
    from fpdf import FPDF
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    unicode_ok = _setup_unicode(pdf)

    # Tiêu đề
    if unicode_ok:
        pdf.set_font("U", "B", 16)
        header = f"Security-App Report — {run_id}"  # tiêu đề đẹp, Unicode OK
    else:
        pdf.set_font("Helvetica", "B", 16)
        header = _ascii_only(f"Security-App Report — {run_id}")  # fallback ASCII
    pdf.cell(0, 10, header, ln=True)

    # Nội dung
    if unicode_ok:
        pdf.set_font("U", size=12)
        def T(x): return str(x)
    else:
        pdf.set_font("Helvetica", size=12)
        def T(x): return _ascii_only(x)

    pdf.cell(0, 8, T("Summary:"), ln=True)
    for k in ("total_rules","all_ok","with_failures","pass_rate",
              "total_commands","commands_ok","commands_failed"):
        pdf.cell(0, 6, T(f"- {k.replace('_', ' ').title()}: {summary.get(k)}"), ln=True)

    pdf.ln(4)
    pdf.cell(0, 8, T("By severity:"), ln=True)
    by = summary.get("by_severity") or {}
    for sev, d in by.items():
        pdf.cell(0, 6, T(f"  {sev}: rules={d.get('rules',0)}, ok={d.get('rules_ok',0)}, "
                         f"cmd_ok={d.get('cmd_ok',0)}, cmd_fail={d.get('cmd_fail',0)}"), ln=True)

    top = summary.get("top_failing_rules") or []
    if top:
        pdf.ln(4)
        pdf.cell(0, 8, T("Top failing rules:"), ln=True)
        for i, t in enumerate(top, start=1):
            title = T((t.get("title") or "")[:100])
            pdf.cell(0, 6, T(f"  {i}. {t.get('id')} [{t.get('severity')}] "
                             f"fail={t.get('cmd_fail')} — {title}"), ln=True)

    den = summary.get("denied_rules") or []
    if den:
        pdf.ln(4)
        pdf.cell(0, 8, T("Denied by safety policy:"), ln=True)
        for r in den[:20]:
            pdf.cell(0, 6, T(f"  {r.get('id')} [{r.get('severity')}] "
                             f"#{r.get('denied')} — {r.get('title')}"), ln=True)

    out = BytesIO()
    pdf.output(out)
    return out.getvalue()

def save_copy_if_configured(data: bytes, run_id: str, ext: str) -> None:
    _maybe_save_copy(data, run_id, ext)
